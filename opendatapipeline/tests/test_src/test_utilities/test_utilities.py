import unittest
from unittest.mock import mock_open, patch, MagicMock
import yaml
from io import BytesIO
from openpyxl import Workbook
# Import your ReadFile class and UtilityException
from src.utilities.utilities import ReadFile
from src.exceptions.exceptions import UtilityException
from src.utilities.utilities import CommonUtils


class TestReadFile(unittest.TestCase):
    def setUp(self):
        self.reader = ReadFile()

    @patch("builtins.open", new_callable=mock_open, read_data='datasources:\n  - driver: mysql\n    connection_string: { "url": "mysql://user:pass@localhost/db" }\n  - driver: postgres\n    connection_string: { "url": "postgres://user:pass@localhost/db" }')
    def test_get_connection_url_dict_success(self, mock_open):
        yml_path = "fake_path.yml"
        engine = "mysql"
        expected_output = { "url": "mysql://user:pass@localhost/db" }

        # Perform the test
        result = self.reader.get_connection_url_dict(yml_path, engine)
        self.assertEqual(result, expected_output)
        mock_open.assert_called_once_with(yml_path, "r")

    @patch("builtins.open", new_callable=mock_open)
    def test_get_connection_url_dict_file_not_found(self, mock_open):
        yml_path = "fake_path.yml"
        engine = "mysql"

        # Simulate file not found by raising an IOError
        mock_open.side_effect = IOError("File not found")

        # Perform the test
        with self.assertRaises(UtilityException) as cm:
            self.reader.get_connection_url_dict(yml_path, engine)
        self.assertEqual(str(cm.exception), "Unable to perform yml reader: File not found")

    @patch("builtins.open", new_callable=mock_open, read_data='invalid_yaml')
    def test_get_connection_url_dict_invalid_yaml(self, mock_open):
        yml_path = "fake_path.yml"
        engine = "mysql"

        # Simulate YAML load error
        with patch("yaml.safe_load", side_effect=yaml.YAMLError("YAML error")):
            with self.assertRaises(UtilityException) as cm:
                self.reader.get_connection_url_dict(yml_path, engine)
            self.assertEqual(str(cm.exception), "Unable to perform yml reader: YAML error")


class TestCommonUtils(unittest.TestCase):
    def setUp(self):
        self.common_utils = CommonUtils()
        self.mock_s3_client = MagicMock()
        self.bucket="dummy-bucket"

    def test_flat_structure(self):
        s3_keys = [
            "file1.csv",
            "file2.txt"
        ]
        expected = [
            {
                "title": "file1.csv",
                "value": "file1.csv",
                "type": "csv",
                "children": None
            },
            {
                "title": "file2.txt",
                "value": "file2.txt",
                "type": "txt",
                "children": None
            }
        ]
        result = self.common_utils.build_tree_from_s3_keys(s3_keys, self.mock_s3_client, self.bucket)
        self.assertCountEqual(result, expected)

    def test_nested_structure(self):
        s3_keys = [
            "folder1/",
            "folder1/file1.csv",
            "folder1/folder2/file2.csv",
            "folder3/",
            "folder3/file3.txt"
        ]
        result = self.common_utils.build_tree_from_s3_keys(s3_keys, self.mock_s3_client, self.bucket)

        # Simple assertions to validate nesting
        self.assertEqual(len(result), 2)  # folder1 and folder3
        folder1 = next(item for item in result if item["title"] == "folder1/")
        self.assertEqual(folder1["type"], "folder")
        self.assertEqual(len(folder1["children"]), 2)

        folder2 = next(child for child in folder1["children"] if child["title"] == "folder2/")
        self.assertEqual(folder2["type"], "folder")
        self.assertEqual(len(folder2["children"]), 1)
        self.assertEqual(folder2["children"][0]["title"], "file2.csv")

    def test_empty_folder_is_removed(self):
        s3_keys = [
            "empty_folder/",               # Empty folder — should be removed
            "non_empty_folder/",
            "non_empty_folder/file.csv"   # Folder with file — should remain
        ]
        result = self.common_utils.build_tree_from_s3_keys(s3_keys, self.mock_s3_client, self.bucket)

        # Should only contain non_empty_folder
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["title"], "non_empty_folder/")
        self.assertEqual(result[0]["type"], "folder")
        self.assertEqual(len(result[0]["children"]), 1)
        self.assertEqual(result[0]["children"][0]["title"], "file.csv")
    
    def test_file_type_detection(self):
        s3_keys = [
            "docs/report.xlsx",
            "data/sample.csv",
            "test_folder/",
            "test_folder/file"  # No extension
        ]
        result = self.common_utils.build_tree_from_s3_keys(s3_keys, self.mock_s3_client, self.bucket)

        def collect_types(nodes, types_dict):
            for node in nodes:
                clean_title = node["title"].rstrip("/")
                types_dict[clean_title] = node["type"]
                if node["type"] == "folder":
                    collect_types(node.get("children", []), types_dict)

        actual_types = {}
        collect_types(result, actual_types)
        expected_types = {
            "docs": "folder",
            "report.xlsx": "xlsx",
            "data": "folder",
            "sample.csv": "csv",
            "test_folder": "folder",
            "file": "file"
        }
        self.assertDictEqual(actual_types, expected_types)

    def test_excel_sheet_values_with_full_path(self):
        s3_keys = [
            "folder/Report.xlsx"
        ]

        # Mock Excel sheet name extraction
        self.common_utils.get_excel_sheetnames_from_s3 = MagicMock(return_value=["Sheet1", "Sheet2"])

        result = self.common_utils.build_tree_from_s3_keys(s3_keys, self.mock_s3_client, self.bucket)

        # Validate structure
        self.assertEqual(len(result), 1)
        report_file = result[0]["children"][0]  # folder -> Report.xlsx
        self.assertEqual(report_file["title"], "Report.xlsx")
        self.assertEqual(report_file["type"], "xlsx")
        self.assertEqual(len(report_file["children"]), 2)

        # Ensure full path is used in sheet value
        expected_values = {"folder/Report.xlsx.Sheet1", "folder/Report.xlsx.Sheet2"}
        actual_values = {sheet["value"] for sheet in report_file["children"]}
        self.assertSetEqual(actual_values, expected_values)

    def test_get_excel_sheetnames_from_s3(self):
        # Create an in-memory Excel file with multiple sheets
        workbook = Workbook()
        workbook.create_sheet("SheetOne")
        workbook.create_sheet("SheetTwo")
        del workbook["Sheet"]  # Remove default sheet
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Mock S3 client to return the Excel file
        mock_s3_client = MagicMock()
        mock_s3_client.get_object.return_value = {"Body": buffer}

        # Call the method
        sheet_names = self.common_utils.get_excel_sheetnames_from_s3(
            s3_client=mock_s3_client,
            bucket="dummy-bucket",
            key="dummy.xlsx"
        )

        # Assert the expected sheet names
        self.assertListEqual(sheet_names, ["SheetOne", "SheetTwo"])

if __name__ == "__main__":
    unittest.main()
