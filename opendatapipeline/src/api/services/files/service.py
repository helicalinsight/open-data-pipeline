from core.mongo.connector import MongoConnector
from ....models.mongo.mongo_files import MongoFiles
from ....models.mongo.mongo_factory import MongoFactory
from ....configurations.api.config import BaseConfig
from ....logger.logger import Logger, logger
from ....utilities.utilities import CommonUtils
from src.cache.cache_factory import get_cache
from src.cache.cache_base import CacheBase
from pymongo.errors import ConnectionFailure, OperationFailure, PyMongoError


import os
import shutil
import uuid
import logging
from openpyxl import load_workbook
from src.api.services.base.service_parent import ServiceParent
from core.mongo.read_connector import ReadMongoConnector

logging.basicConfig(level=logging.DEBUG)

class FileService(ServiceParent):
    """Service class for managing file operations.

    Provides methods for getting a list of files, uploading, deleting, and updating
    file names. It interacts with MongoDB collections for files and users, handling
    file storage and metadata updates.
    """
    def __init__(self,session=None):
        """Constructor method for initializing the FileService.

        :param session: Optional database session object.
        :type session: Session, optional
        """

        super().__init__(session)
        # Initialize MongoFiles instances using the inherited client and session
        self.mongo_files = MongoFiles(self.client, "files", session=self.session)
        self.mongo_users = MongoFiles(self.client, "users", session=self.session)
        
        logging.debug(f"FileService initialized ")
    

    @Logger.generate
    def get_files_list(self, user_id):
        """Retrieves a list of files for the given user.

        Fetches the list of files associated with the specified user ID from the
        MongoDB collection and formats it for response.

        :param user_id: The ID of the user whose files are to be retrieved.
        :type user_id: str
        :return: A dictionary containing the success status and a list of files.
        :rtype: dict
        :raises Exception: If files are not found for the given user.
        """
        status, document = self.mongo_files.get_by_user_id(user_id)
        try:
            if document is None:
                return {"success": True, "filesList": []}, 200
                # raise Exception(f"Files not found with user_id: {user_id}") # pragma: no cover
                # return {
                #     "success": True,
                #     "filesList": []
                # }, 200
            else:
                files_list = [
                    {
                        "_id": file.get("file_id", ""),
                        "alias": file.get("file_name", ""),
                        "fileType": file.get("file_type", ""),
                    }
                    for file in document.get("files", [])
                ]
                logger.info("Successfully got files list")
                return {
                    "success": True,
                    "filesList": files_list
                }, 200
        except Exception as e: # pragma: no cover
            logger.error(f"{e}", exc_info=True)
            self._safe_abort_transaction()  
            return{
                "success": False,
                "filesList": [],
                "message": f"{e}"
            }, 400

    @Logger.generate
    def upload_file(self, user_id, file, file_size):
        """Uploads a file for the given user.

        Saves the file to the specified location and updates or inserts metadata
        into the MongoDB collection.

        :param user_id: The ID of the user who is uploading the file.
        :type user_id: str
        :param file: The file object to be uploaded.
        :type file: werkzeug.datastructures.FileStorage
        :param file_size: The size of the file in bytes.
        :type file_size: int
        :return: A dictionary containing the success status, uploaded file metadata,
                 and a message.
        :rtype: dict
        :raises Exception: If the file size is too large or other upload errors occur.
        """
        try:
            status, user_document = self.mongo_users.get_by_id(user_id)

            if not file:
                raise Exception("Please make sure you selected the files to upload.") # pragma: no cover

            user_folder = os.path.join(BaseConfig.BASE_DIR, BaseConfig.UPLOAD_FOLDER, str(user_id), "sources", "flat_files")
            os.makedirs(user_folder, exist_ok=True)
            file_full_name = file.filename
            file_name, file_type = os.path.splitext(file_full_name)
            file_path = os.path.join(user_folder, file_full_name)
            status, file_exist = self.mongo_files.is_file_exist(user_id, file_name)
            try:
                settings_size = user_document.get("settings").get("files").get("file_size")*1024
            except: # pragma: no cover
                settings_size = 5*1024
            if file_size < settings_size:
                if file_exist :
                    file_id = file_exist["file_id"]
                    file.save(file_path)

                    data = {"metadata": None, "json_path": None, "available": False}
                    self.mongo_files.update_by_file_id(user_id, file_id, data)
                else:
                    file_id = str(uuid.uuid4())
                    data = {
                        "file_id": file_id,
                        "file_name": file_name,
                        "file_type": file_type,
                        "file_path": file_path,
                        "full_name": file_full_name,
                        "metadata": None,
                        "json_path": None,
                        "available": False
                    }
                    self.mongo_files.append_one_by_user_id(user_id, "files", data)

                    file.save(file_path)

                filesUploaded = {"_id": file_id, "alias": file_name, "fileType": file_type}
                logger.info("File uploaded successfully")
                return {
                    "success": True,
                    "filesUploaded": filesUploaded,
                    "message": "File uploaded successfully"
                }, 200
            else:
                logger.error("Please make sure the file size is within 5mb.", exc_info=True)
                raise Exception(f"Please make sure the file size is within {settings_size}mb.")
            
        except OperationFailure as e: # pragma: no cover
            self.session.abort_transaction()
            raise OperationFailure(e)
        except Exception as e: # pragma: no cover
            logger.error(f"An error occurred: {e}", exc_info=True)

            self.session.abort_transaction()
            raise Exception(e) from e


    @Logger.generate
    def delete_file(self, user_id, file_ids):
        """Deletes files for the given user based on file IDs.

        Removes the specified files from the filesystem and the MongoDB collection.

        :param user_id: The ID of the user whose files are to be deleted.
        :type user_id: str
        :param file_ids: List of file IDs to be deleted.
        :type file_ids: list
        :return: A dictionary containing the success status, a message, and a list
                 of failed file IDs.
        :rtype: dict
        :raises Exception: If errors occur during file deletion.
        """
        try:
            overall_success = True
            overall_message = "File deleted successfully."
            failed_files = []

            for file_id in file_ids:
                status, file_info = self.mongo_files.get_by_file_id(user_id, file_id)

                if file_info is None:
                    failed_files.append(file_id)
                    continue

                file_path = file_info.get("file_path", None)
                if file_path is None:
                    failed_files.append(file_id)
                    continue

                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        logger.info("Successfully deleted the files")
                    except PermissionError: # pragma: no cover
                        overall_success = False
                        overall_message = "Failed to delete one or more files due to permission error."
                        logger.error(overall_message, exc_info=True)
                    else:
                        self.mongo_files.delete_by_file_id(user_id, file_id)
                else: # pragma: no cover
                    self.mongo_files.delete_by_file_id(user_id, file_id)

            if overall_success:
                logger.info(overall_message)
                return {"success": True, "message": overall_message, "failed_file_ids": failed_files}, 200
            else: # pragma: no cover
                logger.error("Unable to delete the file.", exc_info=True)
                raise Exception(f"Unable to delete the file with user_id: {user_id} and file_ids: {file_ids}") 
        except Exception as e: # pragma: no cover
            logger.error(f"An error occurred: {e}", exc_info=True)

            self.session.abort_transaction()
            return {"success": False, "message": f"{e}", "failed_file_ids": failed_files}, 500


    @Logger.generate
    def update_file_name(self, user_id, file_id, new_file_name):
        """Updates the name of a file for the given user.

        Renames the file in the filesystem and updates the file metadata in MongoDB.

        :param user_id: The ID of the user whose file name is to be updated.
        :type user_id: str
        :param file_id: The ID of the file to be renamed.
        :type file_id: str
        :param new_file_name: The new name for the file.
        :type new_file_name: str
        :return: A dictionary containing the success status and a message.
        :rtype: dict
        :raises Exception: If errors occur during file renaming.
        """
        try:
            status, file_info = self.mongo_files.get_by_file_id(user_id, file_id)

            if file_info is None: # pragma: no cover
                raise Exception(f"Failed to update file name {new_file_name}")
                # return {"success": False, "message": "Failed to update file name."}, 500

            if new_file_name is None: # pragma: no cover
                raise Exception(f"Failed to update file name {new_file_name}")
                # return {"success": False, "message": "Failed to update file name."}, 500

            original_file_path = file_info.get("file_path")
            original_file_basename, original_file_ext = os.path.splitext(original_file_path)
            updated_file_path = os.path.join(os.path.dirname(original_file_path),
                                            f"{new_file_name}{original_file_ext}")

            if os.path.exists(updated_file_path): # pragma: no cover
                raise Exception(f"Failed to update file name {new_file_name}")
                # return {"success": False, "message": "Failed to update file name."}, 500

            new_data = {"file_name": new_file_name,
                        "file_path": updated_file_path,
                        "full_name": new_file_name + original_file_ext}

            if self.mongo_files.update_by_file_id(user_id, file_id, new_data):
                try:
                    os.rename(original_file_path, updated_file_path)
                    logger.info("Successfully updated file name")
                except PermissionError: # pragma: no cover
                    logger.error("Failed to update file name.", exc_info=True)
                    raise Exception(f"Failed to update file name {new_file_name}") from PermissionError
                    # return {"success": False, "message": "Failed to update file name."}, 500
            else: # pragma: no cover
                raise Exception(f"Failed to update file name {new_file_name}")
                # return {"success": False, "message": "Failed to update file name."}, 500
            logger.info("File name updated successfully.")
            return {"success": True, "message": "File name updated successfully."}, 200
        
        except Exception as e: # pragma: no cover
            logger.error(f"An error occurred: {e}", exc_info=True)
            
            self._safe_abort_transaction()  
            return {"success": False, "message": f"{e}"}, 500
        
    def get_list_catalogs(self, file_id):
        """
        Retrieves a list of catalogs from an Excel file based on the given file ID.
        
        Parameters:
        file_id (str): The ID of the file to retrieve catalogs from.
        
        Returns:
        dict: A response containing the success status and the catalogs list.
        int: HTTP status code (200 for success, 500 for failure).
        """
        try:
            catalogs_list = {"title": "", "value": "", "children": []}
            
            logger.debug(f"Retrieving file information for file_id: {file_id}")
            status, file_info = self.mongo_files.get_by_file_id_only(file_id)

            if not status or file_info is None:
                logger.error("Failed to retrieve file information or file not found.", exc_info=True)
                raise Exception(f"Failed to retrieve file information with file_id: {file_id}")

            file_type = file_info.get("file_type")
            file_name = file_info.get("file_name")
            file_path = file_info.get("file_path")

            catalogs_list['title'] = file_name
            catalogs_list['value'] = file_name
            
            if file_type in ['.xlsx', '.xls']:
                logger.debug(f"Loading workbook for file: {file_path}")
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    catalogs_list["children"].append({
                        "title": sheet_name,
                        "value": f"{file_name}.{sheet_name}"
                    })

            logger.info(f"Successfully retrieved catalogs list for file_id: {file_id}")
            return {"success": True, "dataCatalog": [catalogs_list]}, 200

        except Exception as e:
            logger.error(f"An error occurred while retrieving catalogs list: {e}", exc_info=True)
            self._safe_abort_transaction()  
            return {"success": False, "dataCatalog": [catalogs_list]}, 500



class FileServiceCache:
    """Service class for managing file-related cache operations.

    Provides methods for renaming files in the cache based on their source ID.
    """
    def __init__(self, session=None):
        """Constructor method for initializing the FileServiceCache.

        :param session: Optional database session object.
        :type session: Session, optional
        """
        self.session = session

    def rename_file(self, source_id, new_file_name, chat_id):
        """Renames a file in the cache based on the source ID.

        Updates the file name in the cache collection for the specified source ID.

        :param source_id: The ID of the source for the file to be renamed.
        :type source_id: str
        :param new_file_name: The new name for the file.
        :type new_file_name: str
        :return: A dictionary containing the success status and a message.
        :rtype: dict
        """
        mongo_connector = MongoConnector()
        mongo_client = mongo_connector.client
        cache_db = MongoFactory(mongo_client, "cache", session=self.session)
        chats = MongoFactory(mongo_client, "chats", session=self.session)
        cache: CacheBase = get_cache(session=self.session)
        _, chat_document = chats.get_by_id(chat_id)
        if chat_document is None:
            return {"success": False, "message": "File not found."}, 404
        
        file = cache.get_item(source_id, chat_document.get("user_id"), chat_id)
        if file is None:
            return {"success": False, "message": "File not found."}, 404
        sanatized_name = CommonUtils().replace_special_chars(new_file_name)
        history = chat_document.get("history",[])
        old_alias = file.get("dataframe_alias","")
        history = CommonUtils().replace_alias_in_history(history,old_alias,new_file_name)
        if file is not None:
            cache_db.update_one_by_fields(source_id, chat_id, chat_document.get("user_id"), "dataframe_alias", sanatized_name)
            chats.update_one(chat_id,"history",history)
            return {"success": True, "message": "File renamed successfully."}, 200
